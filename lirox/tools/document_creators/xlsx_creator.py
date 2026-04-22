"""XLSX Creator — Spreadsheet Engine.

Produces a .xlsx workbook with:
  - Topic-aware header fill color
  - Alternating row shading
  - Auto-fitted column widths (capped at 50 characters)
  - Thin cell borders throughout
  - One worksheet per item in *sheets*
"""
from __future__ import annotations

import logging
import os
from typing import Any, Dict, List

from lirox.verify import FileReceipt
from lirox.tools.document_creators.base import ensure_dep, pick_palette, PALETTES
from lirox.safety.audit import log_audit_event

_logger = logging.getLogger("lirox.document_creators.xlsx")


def create_xlsx(path: str, title: str, sheets: List[Dict[str, Any]],
                query: str = "", user_name: str = "", user_expertise: str = "intermediate") -> FileReceipt:
    """Create an Excel workbook with styled headers and data.

    Parameters
    ----------
    path:       Absolute output path.
    title:      Used for the workbook title metadata.
    sheets:     List of dicts with keys ``name``, ``headers``, ``rows``.
    query:      Original user request — used for palette selection.
    user_name:  Recorded in workbook properties.

    Returns
    -------
    FileReceipt with ``ok=True`` and ``verified=True`` on success.
    """
    from pathlib import Path as _Path
    out_path = _Path(path).resolve()
    r = FileReceipt(tool="xlsx_creator", operation="create_xlsx", path=str(out_path))

    # ── PRE-FLIGHT VALIDATION (v1.1 fix) ──
    if not path:
        r.error = "Path cannot be empty"
        return r
    if not title:
        title = "Untitled Spreadsheet"
    if user_name is None:
        user_name = ""
    if sheets is None or not isinstance(sheets, list):
        sheets = [{"name": "Sheet1", "headers": ["Data"], "rows": []}]
    if not sheets:
        sheets = [{"name": "Sheet1", "headers": ["Data"], "rows": []}]

    try:
        ensure_dep("openpyxl")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        out_dir = out_path.parent
        out_dir.mkdir(parents=True, exist_ok=True)
        if not os.access(str(out_dir), os.W_OK):
            r.error = f"Output directory is not writable: {out_dir}"
            return r

        palette_name = pick_palette(query or title, title, user_expertise=user_expertise)
        pal = PALETTES[palette_name]

        wb = Workbook()
        wb.remove(wb.active)
        if hasattr(wb, "properties"):
            wb.properties.creator = user_name or "Generated Document"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color=pal["primary"], end_color=pal["primary"],
            fill_type="solid")
        alt_fill = PatternFill(
            start_color=pal["secondary"], end_color=pal["secondary"],
            fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"))

        total_rows = 0
        for sd in sheets:
            ws      = wb.create_sheet(title=str(sd.get("name", "Sheet"))[:31])
            headers = sd.get("headers", [])
            rows    = sd.get("rows", [])

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=str(h))
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border    = thin_border

            for ri, row_data in enumerate(rows, 2):
                for ci, val in enumerate(row_data, 1):
                    safe_val = str(val) if val is not None else ""
                    cell = ws.cell(row=ri, column=ci, value=safe_val)
                    cell.border = thin_border
                    if ri % 2 == 0:
                        cell.fill = alt_fill
                total_rows += 1

            for col in ws.columns:
                max_len = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        wb.save(str(out_path))

        if out_path.exists():
            r.ok = True
            r.verified = True
            r.bytes_written = out_path.stat().st_size
            r.message = (
                f"Created Excel: {out_path} "
                f"({r.bytes_written:,} bytes, {len(sheets)} sheet(s), "
                f"{total_rows} data rows)"
            )
            r.details["sheet_count"] = len(sheets)
            r.details["row_count"]   = total_rows
            r.details["palette"]     = palette_name
        else:
            r.error = "Excel build completed but file not found on disk"
        log_audit_event(
            "document_create_xlsx",
            str(out_path),
            status="ok" if (r.ok and r.verified) else "error",
            detail=r.message or r.error,
        )
        return r
    except Exception as e:
        r.error = f"Excel creation error: {e}"
        log_audit_event("document_create_xlsx", str(out_path), status="error", detail=r.error)
        return r
